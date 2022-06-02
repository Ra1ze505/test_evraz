from typing import Optional
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import IntegrityError
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from rest_framework.generics import get_object_or_404


from ZMK.models import ZMK, RTC, RTCObject


COLUMN_NAMES = {
    '№': 'number',
    'Дата': 'date',
    'Вес(тн)': 'weight',
    'Статус': 'status',
    'Дата выгрузки': 'unloading_date',
    '№ УПД': 'upd'
}


class ZMKTableReader:

    def __init__(self, cell: Cell, zmk: ZMK, next_cell: Optional[Cell]):
        self.cell = cell
        self.zmk = zmk
        self.start_row = cell.row
        self.start_column = cell.column
        self.end_column = next_cell.column - 1 if next_cell else None
        self.rtcs: list[dict] = []
        self.rtc_objects: list[dict] = []

    def _get_columns(self, ws):
        columns = {'objects': []}
        for row in ws.iter_rows(min_row=self.start_row + 2, max_row=self.start_row + 2, min_col=self.start_column,
                                max_col=self.end_column):
            for index, cell in enumerate(row):
                column_name = COLUMN_NAMES.get(cell.value)
                if column_name is not None:
                    columns[COLUMN_NAMES.get(cell.value, None)] = index
                else:
                    columns['objects'].append(index)
        return columns

    def read_rows(self, ws):
        columns = self._get_columns(ws)

        for row in ws.iter_rows(min_row=self.start_row + 3, min_col=self.start_column, max_col=self.end_column):
            date = self._get_value(row, columns.get('date'))
            weight = self._get_value(row, columns.get('weight'))
            status = self._get_value(row, columns.get('status'))
            upd = self._get_value(row, columns.get('upd'))
            unloading_date = self._get_value(row, columns.get('unloading_date'))
            if date is None:
                break
            if unloading_date is None:
                continue
            rtc = dict(zmk=self.zmk, date=date, weight=weight, status=status, upd=upd, unloading_date=unloading_date)
            rtc_objects = [dict(object_weight=self._get_value(row, index)) for index in columns.get('objects')]
            self.rtcs.append(rtc)
            self.rtc_objects.append(rtc_objects)

        return self.rtcs, self.rtc_objects

    def _get_value(self, row, column_index):
        if column_index is not None:
            return row[column_index].value
        return None


class ImportService:

    def __init__(self, file: InMemoryUploadedFile):
        self.file = file
        self.tables: list[ZMKTableReader] = []
        self.ws = load_workbook(filename=self.file, data_only=True).active

    def import_data(self):
        """Import data from excel file"""
        self._setup_tables()
        for table in self.tables:
            _rtcs, _rtc_objects = table.read_rows(self.ws)
            # print(_rtc_objects)
            self.bulk_create_or_update_data(_rtcs, _rtc_objects)

    def bulk_create_or_update_data(self, rtcs_list: list[dict], rtc_objects_list: list[list[dict]]):
        """Bulk create or update data"""
        rtcs = [RTC(**rtc) for rtc in rtcs_list]
        bulk_create_objects = []
        bulk_update_objects = []

        for index, rtc in enumerate(rtcs):
            try:
                rtc.save()
                for object in rtc_objects_list[index]:
                    object['rtc_id'] = rtc.pk
                    if object['object_weight'] is None:
                        continue
                    bulk_create_objects.append(RTCObject(**object))
            except IntegrityError:
                bulk_update_objects.append(rtc)

        for obj in bulk_update_objects:
            obj_in_db = RTC.objects.get(zmk_id=obj.zmk_id,
                                        date=obj.date,
                                        unloading_date=obj.unloading_date,
                                        weight=obj.weight)
            obj.pk = obj_in_db.pk

        RTC.objects.bulk_update(bulk_update_objects, ['weight', 'status', 'upd', 'unloading_date'])
        RTCObject.objects.bulk_create(bulk_create_objects, ignore_conflicts=True)

    def _get_zmk(self, **kwargs):
        return get_object_or_404(ZMK, **kwargs)

    def _setup_tables(self) -> list[ZMKTableReader]:
        """Search for tables in file"""
        data = []
        for row in self.ws.rows:
            for cell in row:
                if cell.value is not None:
                    data.append(cell)
            if data:
                break

        for index, cell in enumerate(data):
            if len(data) > index + 1:
                self.tables.append(ZMKTableReader(cell, self._get_zmk(title=cell.value), data[index + 1]))
            else:
                self.tables.append(ZMKTableReader(cell, self._get_zmk(title=cell.value), None))
